# -*- coding: utf-8 -*-
"""
Created on Fri Mar 27 14:15:19 2020

@author: connor.folkman
"""
import pandas as pd
from pandas import DataFrame
import glob
import os.path, time
from openpyxl import load_workbook

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    
#creates empty list to be populated with peak tensile test data and date created from each file for the hub
peak_tensile_data_hub = []
file_creation_date_hub = []
WO_number_hub = []
part_number_hub = []

#creates empty list to be populated with peak tensile test data and date created from each file for the tip
peak_tensile_data_tip = []
file_creation_date_tip = []
WO_number_tip = []
part_number_tip = []

#creates empty list to be populated with peak tensile test data and date created from each file for the mb
peak_tensile_data_mb = []
file_creation_date_mb = []
WO_number_mb = []
part_number_mb = []

#created empty data frames to put the data into
df = DataFrame()
df_2 = DataFrame()
completed_WO = False

#opens directory where tensile data is saved and reads from each file to find the peak force from colume 'Load' for the hub
for filepath in glob.iglob('//merit.com/Merit_Shares/South_Jordan/public/MAESTRO/MAESTRO/TENSILE TESTER DATA FILES/HUB TENSILE/*.xlsx'):
    df_2 = pd.read_excel(io='//merit.com/Merit_Shares/South_Jordan/public/MAESTRO/MAESTRO/ENGINEERING/TENSILE DATA/TENSILE DATA SPREADSHEET.xlsx', sheet_name='Hub Tensile')
    completed_WO_list = df_2['WO:'].to_list()
    i = 0
    while i < len(completed_WO_list):
        if filepath[99:107] == completed_WO_list[i]:
            completed_WO = True
            break
        else: completed_WO = False
        i += 1
    if completed_WO == False:
        df = pd.read_excel(filepath)
        file_creation_date_hub.append(time.strftime('%m/%d/%Y', time.localtime(os.path.getmtime(filepath))))
        WO_number_hub.append(filepath[99:107])
        part_number_hub.append(filepath[108:-5])
        try: 
            if df['Load [N]'].max() > 0:
                peak_tensile_data_hub.append(df['Load [N]'].max())
        except:
            try: 
                if df['Load [lbf]'].max() > 0:
                    peak_tensile_data_hub.append(df['Load [lbF]'].max()/0.73756)
            except:
                peak_tensile_data_hub.append(0)
    else: continue
  
#opens directory where tensile data is saved and reads from each file to find the peak force from colume 'Load' for the tip
for filepath in glob.iglob('//merit.com/Merit_Shares/South_Jordan/public/MAESTRO/MAESTRO/TENSILE TESTER DATA FILES/TIP TENSILE/*.xlsx'):
    df_2 = pd.read_excel(io='//merit.com/Merit_Shares/South_Jordan/public/MAESTRO/MAESTRO/ENGINEERING/TENSILE DATA/TENSILE DATA SPREADSHEET.xlsx', sheet_name='Tip Tensile')
    completed_WO_list = df_2['WO:'].to_list()
    i = 0
    while i < len(completed_WO_list):
        if filepath[99:107] == completed_WO_list[i]:
            completed_WO = True
            break
        else: completed_WO = False
        i += 1
    if completed_WO == False:
        df = pd.read_excel(filepath)
        file_creation_date_tip.append(time.strftime('%m/%d/%Y', time.localtime(os.path.getmtime(filepath))))
        WO_number_tip.append(filepath[99:107])
        part_number_tip.append(filepath[108:-5])
        try: 
            if df['Load [N]'].max() > 0:
                peak_tensile_data_tip.append(df['Load [N]'].max())
        except:
            try: 
                if df['Load [lbf]'].max() > 0:
                    peak_tensile_data_tip.append(df['Load [lbF]'].max()/0.73756)
            except:
                peak_tensile_data_tip.append(0)
    else: continue
            
#opens directory where tensile data is saved and reads from each file to find the peak force from colume 'Load' for the mb
for filepath in glob.glob('//merit.com/Merit_Shares/South_Jordan/public/MAESTRO/MAESTRO/TENSILE TESTER DATA FILES/MB TENSILE/*.xlsx'):
    df_2 = pd.read_excel(io='//merit.com/Merit_Shares/South_Jordan/public/MAESTRO/MAESTRO/ENGINEERING/TENSILE DATA/TENSILE DATA SPREADSHEET.xlsx', sheet_name='MB Tensile')
    completed_WO_list = df_2['WO:'].to_list()
    i = 0
    while i < len(completed_WO_list):
        if filepath[98:106] == completed_WO_list[i]:
            completed_WO = True
            break
        else: completed_WO = False
        i += 1
    if completed_WO == False:
        df = pd.read_excel(filepath)
        file_creation_date_mb.append(time.strftime('%m/%d/%Y', time.localtime(os.path.getmtime(filepath))))
        WO_number_mb.append(filepath[98:106])
        part_number_mb.append(filepath[107:-5])
        try: 
            if df['Load [lbf]'].max() > 0:
                peak_tensile_data_mb.append(df['Load [lbF]'].max())
        except:
            try: 
                if df['Load [N]'].max() > 0:
                    peak_tensile_data_mb.append(df['Load [N]'].max()*0.73756)
            except:
                peak_tensile_data_mb.append(0)
    else: continue

#creates new dataframe to load data into, sorting by date file was created
df_mb = pd.DataFrame({'Date Test Performed for MB' : file_creation_date_mb, 'WO:' : WO_number_mb, 'Part Number:' : part_number_mb, 'Peak Tensile [lbf] for MB' : peak_tensile_data_mb})
df_hub = pd.DataFrame({'Date Test Performed for Hub' : file_creation_date_hub, 'WO:' : WO_number_hub, 'Part Number:' : part_number_hub, 'Peak Tensile [N] for Hub' : peak_tensile_data_hub}) 
df_tip = pd.DataFrame({'Date Test Performed for Tip' : file_creation_date_tip, 'WO:' : WO_number_tip, 'Part Number:' : part_number_tip, 'Peak Tensile [N] for Tip' : peak_tensile_data_tip})

#opens excel file and appends new tensile data to end of spreadsheet
append_df_to_excel('//merit.com/Merit_Shares/South_Jordan/public/MAESTRO/MAESTRO/ENGINEERING/TENSILE DATA/TENSILE DATA SPREADSHEET.xlsx', df_hub, sheet_name='Hub Tensile', index=False)
append_df_to_excel('//merit.com/Merit_Shares/South_Jordan/public/MAESTRO/MAESTRO/ENGINEERING/TENSILE DATA/TENSILE DATA SPREADSHEET.xlsx', df_tip, sheet_name='Tip Tensile', index=False)
append_df_to_excel('//merit.com/Merit_Shares/South_Jordan/public/MAESTRO/MAESTRO/ENGINEERING/TENSILE DATA/TENSILE DATA SPREADSHEET.xlsx', df_mb, sheet_name='MB Tensile', index=False)
